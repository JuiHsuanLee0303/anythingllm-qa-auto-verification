import pandas as pd
from typing import List, Dict, Tuple
import openpyxl
import os
from logger import Logger # Assuming Logger is in a file named logger.py

class ExcelHandler:
    def __init__(self, file_path: str, logger: Logger):
        """
        Initialize ExcelHandler, loading the workbook into memory.
        """
        self.file_path = file_path
        self.logger = logger
        
        # 檢查檔案是否存在
        if not os.path.exists(file_path):
            error_msg = f"[ERROR] Excel 檔案不存在: {file_path}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        # 檢查檔案副檔名
        file_extension = os.path.splitext(file_path)[1].lower()
        supported_extensions = ['.xlsx', '.xlsm', '.xltx', '.xltm']
        
        if file_extension not in supported_extensions:
            error_msg = f"[ERROR] 不支援的檔案格式: {file_extension}。支援的格式: {', '.join(supported_extensions)}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        
        # 檢查檔案大小
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            error_msg = f"[ERROR] Excel 檔案是空的: {file_path}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        
        # 嘗試載入檔案
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.logger.info(f"[INFO] 成功載入 Excel 檔案: {file_path} (大小: {file_size} bytes)")
        except openpyxl.utils.exceptions.InvalidFileException as e:
            error_msg = f"[ERROR] 檔案格式錯誤: {file_path}。請確認檔案是有效的 Excel 檔案 (.xlsx, .xlsm, .xltx, .xltm)，並且可以用 Excel 開啟。錯誤詳情: {str(e)}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        except FileNotFoundError:
            error_msg = f"[ERROR] Excel 檔案未找到: {file_path}"
            self.logger.error(error_msg)
            raise
        except Exception as e:
            error_msg = f"[ERROR] 載入 Excel 檔案時發生錯誤: {e}"
            self.logger.error(error_msg, exc_info=True)
            raise

    def get_all_sheets(self) -> List[str]:
        """
        Get all sheet names from the Excel file.
        
        Returns:
            List[str]: List of sheet names
        """
        return self.workbook.sheetnames
    
    def get_qa_pairs(self, sheet_name: str) -> List[Tuple[str, str, int]]:
        """
        Extract Q&A pairs from a specific sheet.
        First column is treated as questions, second column as answers.
        Includes the first row of data even if it's not a header.
        
        Args:
            sheet_name (str): Name of the sheet to process
            
        Returns:
            List[Tuple[str, str, int]]: List of (question, answer, original_row_index) pairs
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
            
            # Ensure there are at least 2 columns
            if len(df.columns) < 2:
                self.logger.warning(f"[WARNING] 工作表 '{sheet_name}' 的欄數少於 2，將被跳過。")
                return []
            
            # Get first two columns
            questions = df.iloc[:, 0].astype(str)
            answers = df.iloc[:, 1].astype(str)
            
            # Filter out empty rows and keep original row indices
            qa_pairs = []
            for i, (q, a) in enumerate(zip(questions, answers)):
                if q.strip() and a.strip():
                    qa_pairs.append((q.strip(), a.strip(), i))
            
            return qa_pairs
        except Exception as e:
            self.logger.error(f"[ERROR] 處理工作表 '{sheet_name}' 時發生錯誤: {e}", exc_info=True)
            return []
    
    def get_all_qa_pairs(self) -> Dict[str, List[Tuple[str, str]]]:
        """
        Get Q&A pairs from all sheets in the Excel file.
        """
        # We need to use pandas to get all sheets initially.
        try:
            excel_file = pd.ExcelFile(self.file_path)
            sheet_names = excel_file.sheet_names

            result = {}
            for sheet_name in sheet_names:
                result[sheet_name] = self.get_qa_pairs(sheet_name)
            return result
        except Exception as e:
            self.logger.error(f"[ERROR] 從 Excel 檔案讀取所有工作表時發生錯誤: {e}", exc_info=True)
            return {}

    def write_llm_response(self, sheet_name: str, row_index: int, llm_response: str) -> None:
        """
        Write LLM response to the third column. Does not save immediately.
        row_index is the original row index from the Excel file (0-based).
        """
        try:
            sheet = self.workbook[sheet_name]
            sheet.cell(row=row_index + 1, column=3, value=llm_response)
        except Exception as e:
            self.logger.error(f"[ERROR] 寫入 LLM 回應時發生錯誤: {e}", exc_info=True)

    def write_similarity_scores(self, sheet_name: str, row_index: int, similarity_scores: dict) -> None:
        """
        Write similarity scores. Does not save immediately.
        row_index is the original row index from the Excel file (0-based).
        """
        try:
            sheet = self.workbook[sheet_name]
            
            # 寫入所有可用的相似度分數
            col = 4  # 從第4欄開始
            for key, value in similarity_scores.items():
                sheet.cell(row=row_index + 1, column=col, value=value)
                col += 1
                
        except Exception as e:
            self.logger.error(f"[ERROR] 寫入相似度分數時發生錯誤: {e}", exc_info=True)

    def save_workbook(self, output_path: str):
        """
        Saves the workbook to a new file path.
        """
        try:
            self.workbook.save(output_path)
            self.logger.info(f"[SUCCESS] Excel 檔案成功儲存至: {output_path}")
        except Exception as e:
            self.logger.error(f"[ERROR] 儲存 Excel 檔案至 '{output_path}' 時發生錯誤: {e}", exc_info=True)
    
    def get_total_qa_pairs(self) -> int:
        """
        計算所有工作表中有效的問答對總數。
        """
        total = 0
        all_pairs = self.get_all_qa_pairs()
        for sheet_name, pairs in all_pairs.items():
            total += len(pairs)
        return total

def demo():
    """
    Demo function to show how to use ExcelHandler
    """
    # Example usage
    try:
        # Create an instance of ExcelHandler
        handler = ExcelHandler("techman_robot.xlsx")
        
        # Get all sheet names
        sheets = handler.get_all_sheets()
        print("\n" + "="*50)
        print(f"[INFO] 找到的工作表: {', '.join(sheets)}")
        print("="*50)
        
        # Process each sheet
        for sheet_name in sheets:
            print(f"\n[INFO] 工作表: {sheet_name}")
            print("-"*50)
            qa_pairs = handler.get_qa_pairs(sheet_name)
            
            print(f"[INFO] 找到 {len(qa_pairs)} 個問答對:")
            for i, (question, answer) in enumerate(qa_pairs, 1):
                print(f"\nQ{i}. {question}")
                print(f"A{i}.\n{answer}")
                print("-"*30)
                
    except FileNotFoundError:
        print("\n[ERROR] 錯誤: techman_robot.xlsx 檔案不存在")
    except Exception as e:
        print(f"\n[ERROR] 錯誤: {str(e)}")

if __name__ == "__main__":
    demo()
