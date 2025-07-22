import os
import uuid
import argparse
import threading
import queue
import shutil
import time
import pandas as pd
import json
import requests
import re
from flask import Flask, render_template, jsonify, request, Response, send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# 匯入重構後的核心邏輯
from config import Config
from logger import get_logger, Logger
from main import run_verification, run_single_verification
from excel_handler import ExcelHandler

# --- App State & Initialization ---

# 建立上傳資料夾
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 建立輸出資料夾
OUTPUT_FOLDER = 'output'
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB

# 用於追蹤背景任務的狀態
tasks = {}

# 建立一個給 Flask 應用本身使用的 logger
app_logger = get_logger("FlaskWebApp")

def cleanup_expired_tasks():
    """清理過期的任務（超過1小時）"""
    current_time = time.time()
    expired_tasks = []
    for task_id, task_info in tasks.items():
        created_time = task_info.get('created_time', 0)
        if current_time - created_time > 3600:  # 1小時過期
            expired_tasks.append(task_id)
    
    for task_id in expired_tasks:
        try:
            del tasks[task_id]
            app_logger.info(f"已清理過期任務: {task_id}")
        except KeyError:
            pass  # 任務可能已被其他進程清理

# --- Helper Function for Threading ---

def run_verification_threaded(task_id: str, config: Config, logger, args: argparse.Namespace, advanced_options: dict):
    """在背景執行緒中運行的包裝函式"""
    log_queue = tasks[task_id]['queue']
    try:
        # --- Override config with advanced options from frontend ---
        if advanced_options.get('api_url'):
            config.api.base_url = advanced_options['api_url']
            logger.info(f"使用前端設定的 API URL: {config.api.base_url}")
        
        if advanced_options.get('api_key'):
            config.api.api_key = advanced_options['api_key']
            logger.info("使用前端提供的 API Key。")

        if advanced_options.get('model'):
            config.workspace.model = advanced_options['model']
            logger.info(f"使用前端設定的 LLM 模型: {config.workspace.model}")

        if advanced_options.get('similarity_threshold'):
            config.analyzer.similarity_threshold = float(advanced_options['similarity_threshold'])
            logger.info(f"使用前端設定的相似度閾值: {config.analyzer.similarity_threshold}")
        # --- End of config override ---

        tasks[task_id]['status'] = 'running'
        logger.info(f"Task {task_id}: 驗證流程開始。")
        run_verification(config, logger, args, web_mode=True)
        tasks[task_id]['status'] = 'completed'

        # 將處理完的 Excel 檔案複製到輸出目錄
        final_excel_path = os.path.join(args.output, os.path.basename(args.excel))
        shutil.copy(args.excel, final_excel_path)
        logger.info(f"結果 Excel 已複製到: {final_excel_path}")
        
        logger.info(f"Task {task_id}: 驗證流程成功完成。")
    except Exception as e:
        logger.error(f"Task {task_id}: 驗證流程發生錯誤: {e}", exc_info=True)
        tasks[task_id]['status'] = 'error'
    finally:
        # 發送結束信號
        log_queue.put("<<TASK_DONE>>")

@app.route('/api/verify', methods=['POST'])
def verify():
    """處理 Excel 檔案驗證請求"""
    try:
        # 檢查必要欄位
        if 'workspace' not in request.form:
            return jsonify({"error": "缺少工作區名稱"}), 400
        
        if 'excel_file' not in request.files:
            return jsonify({"error": "缺少 Excel 檔案"}), 400
        
        workspace = request.form['workspace']
        excel_file = request.files['excel_file']
        
        if excel_file.filename == '':
            return jsonify({"error": "未選擇檔案"}), 400
        
        # 生成任務 ID
        task_id = str(uuid.uuid4())
        
        # 建立任務目錄
        task_dir = os.path.join(OUTPUT_FOLDER, task_id)
        os.makedirs(task_dir, exist_ok=True)
        
        # 儲存上傳的檔案
        filename = secure_filename(excel_file.filename)
        excel_path = os.path.join(UPLOAD_FOLDER, f"{task_id}_{filename}")
        excel_file.save(excel_path)
        
        # 建立任務狀態追蹤
        tasks[task_id] = {
            'status': 'pending',
            'queue': queue.Queue(),
            'created_time': time.time()
        }
        
        # 載入配置
        config = Config.load()
        
        # 解析進階選項
        advanced_options = {}
        if 'api_url' in request.form and request.form['api_url']:
            advanced_options['api_url'] = request.form['api_url']
        if 'api_key' in request.form and request.form['api_key']:
            advanced_options['api_key'] = request.form['api_key']
        if 'model' in request.form and request.form['model']:
            advanced_options['model'] = request.form['model']
        if 'similarity_threshold' in request.form and request.form['similarity_threshold']:
            advanced_options['similarity_threshold'] = request.form['similarity_threshold']
        
        # 建立參數物件
        args = argparse.Namespace()
        args.workspace = workspace
        args.excel = excel_path
        args.output = task_dir
        args.verbose = True
        args.directory = None  # 批次驗證不需要上傳文件目錄
        
        # 建立任務專用的 logger
        task_logger = Logger(task_id, log_queue=tasks[task_id]['queue'])
        
        # 在背景執行緒中執行驗證
        thread = threading.Thread(
            target=run_verification_threaded,
            args=(task_id, config, task_logger, args, advanced_options)
        )
        thread.daemon = True
        thread.start()
        
        app_logger.info(f"Task {task_id}: 已啟動 Excel 驗證任務")
        return jsonify({"task_id": task_id, "message": "驗證任務已啟動"})
        
    except Exception as e:
        app_logger.error(f"驗證請求處理錯誤: {e}", exc_info=True)
        return jsonify({"error": f"處理請求時發生錯誤: {str(e)}"}), 500

@app.route('/api/verify_single', methods=['POST'])
def verify_single():
    """處理單筆文字驗證請求"""
    try:
        # 檢查必要欄位
        if 'workspace' not in request.form:
            return jsonify({"error": "缺少工作區名稱"}), 400
        
        if 'single_question' not in request.form or not request.form['single_question'].strip():
            return jsonify({"error": "缺少問題內容"}), 400
        
        if 'single_answer' not in request.form or not request.form['single_answer'].strip():
            return jsonify({"error": "缺少標準答案"}), 400
        
        workspace = request.form['workspace']
        question = request.form['single_question'].strip()
        standard_answer = request.form['single_answer'].strip()
        
        # 生成任務 ID
        task_id = str(uuid.uuid4())
        
        # 建立任務目錄
        task_dir = os.path.join(OUTPUT_FOLDER, task_id)
        os.makedirs(task_dir, exist_ok=True)
        
        # 建立任務狀態追蹤
        tasks[task_id] = {
            'status': 'pending',
            'queue': queue.Queue(),
            'created_time': time.time()
        }
        
        # 載入配置
        config = Config.load()
        
        # 解析進階選項
        advanced_options = {}
        if 'api_url' in request.form and request.form['api_url']:
            advanced_options['api_url'] = request.form['api_url']
        if 'api_key' in request.form and request.form['api_key']:
            advanced_options['api_key'] = request.form['api_key']
        if 'model' in request.form and request.form['model']:
            advanced_options['model'] = request.form['model']
        if 'similarity_threshold' in request.form and request.form['similarity_threshold']:
            advanced_options['similarity_threshold'] = request.form['similarity_threshold']
        
        # 建立任務專用的 logger
        task_logger = Logger(task_id, log_queue=tasks[task_id]['queue'])
        
        # 在背景執行緒中執行單筆驗證
        thread = threading.Thread(
            target=run_single_verification_threaded,
            args=(task_id, config, task_logger, workspace, question, standard_answer, advanced_options)
        )
        thread.daemon = True
        thread.start()
        
        app_logger.info(f"Task {task_id}: 已啟動單筆文字驗證任務")
        return jsonify({"task_id": task_id, "message": "驗證任務已啟動"})
        
    except Exception as e:
        app_logger.error(f"單筆驗證請求處理錯誤: {e}", exc_info=True)
        return jsonify({"error": f"處理請求時發生錯誤: {str(e)}"}), 500

def run_single_verification_threaded(task_id: str, config: Config, logger, workspace: str, question: str, standard_answer: str, advanced_options: dict):
    """在背景執行緒中運行的單筆驗證包裝函式"""
    log_queue = tasks[task_id]['queue']
    try:
        # --- Override config with advanced options from frontend ---
        if advanced_options.get('api_url'):
            config.api.base_url = advanced_options['api_url']
            logger.info(f"使用前端設定的 API URL: {config.api.base_url}")
        
        if advanced_options.get('api_key'):
            config.api.api_key = advanced_options['api_key']
            logger.info("使用前端提供的 API Key。")

        if advanced_options.get('model'):
            config.workspace.model = advanced_options['model']
            logger.info(f"使用前端設定的 LLM 模型: {config.workspace.model}")

        if advanced_options.get('similarity_threshold'):
            config.analyzer.similarity_threshold = float(advanced_options['similarity_threshold'])
            logger.info(f"使用前端設定的相似度閾值: {config.analyzer.similarity_threshold}")
        # --- End of config override ---

        tasks[task_id]['status'] = 'running'
        logger.info(f"Task {task_id}: 單筆驗證流程開始。")
        
        # 建立臨時的 Excel 檔案
        temp_excel_path = os.path.join(UPLOAD_FOLDER, f"{task_id}_single_verification.xlsx")
        create_single_verification_excel(temp_excel_path, question, standard_answer)
        
        # 建立參數物件
        args = argparse.Namespace()
        args.workspace = workspace
        args.excel = temp_excel_path
        args.output = os.path.join(OUTPUT_FOLDER, task_id)
        args.verbose = True
        
        # 執行驗證並獲取結果
        result = run_single_verification_with_result(config, logger, args, question, standard_answer, web_mode=True)
        
        # 儲存結果到任務狀態中
        if result:
            tasks[task_id]['single_result'] = result
            tasks[task_id]['status'] = 'completed'
            logger.info(f"Task {task_id}: 單筆驗證流程成功完成。")
            logger.info(f"Task {task_id}: 結果已儲存到記憶體中: {result}")
        else:
            tasks[task_id]['status'] = 'error'
            logger.error(f"Task {task_id}: 單筆驗證流程失敗。")
            
    except Exception as e:
        logger.error(f"Task {task_id}: 單筆驗證流程發生錯誤: {e}", exc_info=True)
        tasks[task_id]['status'] = 'error'
    finally:
        # 發送結束信號
        log_queue.put("<<TASK_DONE>>")

def run_single_verification_with_result(config: Config, logger, args: argparse.Namespace, question: str, standard_answer: str, web_mode: bool = False):
    """
    執行單筆文字驗證流程並回傳結果
    
    Args:
        config (Config): 系統配置物件
        logger (Logger): 日誌記錄器實例
        args (argparse.Namespace): 命令列參數
        question (str): 問題內容
        standard_answer (str): 標準答案
        web_mode (bool): 是否為 Web 模式，影響日誌和進度條顯示
        
    Returns:
        dict: 包含驗證結果的字典，如果失敗則回傳 None
    """
    logger.info("🚀 單筆文字驗證系統啟動")
    logger.info(f"工作區: {args.workspace}", progress=5, status="初始化...")

    # 直接在這裡實現單筆驗證邏輯，避免 Excel 檔案讀取問題
    from main import QAVerificationSystem
    
    system = QAVerificationSystem(config, logger)
    
    # 1. 驗證 API 金鑰
    if not system.validate_api_key():
        logger.error("❌ API 金鑰無效，終止程序。")
        return None

    # 2. 獲取或創建工作區
    workspace_slug = system.get_workspace_slug(args.workspace)
    if not workspace_slug:
        workspace_slug = system.create_workspace(args.workspace)
    
    if not workspace_slug:
        logger.error("❌ 無法獲取或創建工作區，終止程序。")
        return None
        
    logger.info(f"✅ 工作區 '{args.workspace}' (slug: {workspace_slug}) 已就緒", progress=30, status="正在發送問題到 LLM...")

    # 3. 發送問題到 AnythingLLM 獲取回答
    try:
        logger.info("正在發送問題到 LLM...", progress=50, status="獲取 LLM 回答...")
        
        response = system.send_chat_message(workspace_slug, question)
        if response and 'textResponse' in response:
            llm_response = response['textResponse']
            # 清理<think></think>之間的文字
            cleaned_llm_response = re.sub(r'<think>.*?</think>', '', llm_response, flags=re.DOTALL).strip()
            
            logger.info("正在計算相似度分數...", progress=70, status="計算相似度...")
            
            similarity_scores = system.similarity_analyzer.calculate_similarity(
                cleaned_llm_response, standard_answer
            )
            
            logger.info(f"✅ 相似度分析完成", progress=80, status="生成報告...")
            
            # 4. 生成總結圖表
            output_dir = args.output
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                
            try:
                system.similarity_analyzer.generate_charts(
                    [similarity_scores], 
                    output_dir
                )
                logger.info(f"📊 分析報告已生成於 '{output_dir}' 目錄。")
                
            except Exception as e:
                logger.error(f"❌ 生成圖表時發生錯誤: {e}", exc_info=True)
            
            # 5. 儲存包含結果的 Excel 檔案（保持原有功能）
            output_excel_path = os.path.join(output_dir, os.path.basename(args.excel))
            try:
                excel_handler = ExcelHandler(args.excel, logger)
                # 將結果寫入 Excel
                excel_handler.write_llm_response("單筆驗證", 0, cleaned_llm_response)
                excel_handler.write_similarity_scores("單筆驗證", 0, similarity_scores)
                excel_handler.save_workbook(output_excel_path)
                logger.info(f"💾 更新後的 Excel 檔案已儲存至: {output_excel_path}", progress=100, status="完成")
            except Exception as e:
                logger.error(f"❌ 儲存 Excel 檔案時發生錯誤: {e}", exc_info=True)

            logger.info("🎉 單筆文字驗證流程全部完成！")
            
            # 直接回傳結果，不從 Excel 檔案讀取
            return {
                "question": question,
                "standard_answer": standard_answer,
                "llm_response": cleaned_llm_response,
                "similarity_scores": similarity_scores
            }
            
        else:
            logger.error("❌ 無法從 LLM 獲取回答")
            return None
            
    except Exception as e:
        logger.error(f"❌ 相似度分析時發生錯誤: {e}", exc_info=True)
        return None

def create_single_verification_excel(excel_path: str, question: str, standard_answer: str):
    """建立單筆驗證用的 Excel 檔案"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "單筆驗證"
    
    # 設定標題 - 包含所有欄位
    headers = ['問題', '標準答案', 'LLM 回答', 'BERT Score', '餘弦相似度']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 設定資料
    ws.cell(row=2, column=1, value=question)
    ws.cell(row=2, column=2, value=standard_answer)
    
    # 調整欄寬
    for col in range(1, 6):  # 5 個欄位
        ws.column_dimensions[chr(64 + col)].width = 40
    
    # 儲存檔案
    wb.save(excel_path)

# --- Routes ---

@app.route('/')
def index():
    """提供主頁面並附上快取清除參數"""
    config = Config.load()
    default_api_url = config.api.base_url
    default_api_key = config.api.api_key or ""  # 確保不是 None
    return render_template(
        'index.html', 
        version=time.time(),
        default_api_url=default_api_url,
        default_api_key=default_api_key
    )

@app.route('/stream/<task_id>')
def stream(task_id: str):
    """此端點為客戶端提供 Server-Sent Events (SSE)"""
    # 清理過期任務
    cleanup_expired_tasks()
    
    log_queue = tasks.get(task_id, {}).get('queue')
    if not log_queue:
        return Response("錯誤：找不到任務佇列或任務不存在。", status=404)

    def event_stream():
        start_time = time.time()
        last_heartbeat = start_time
        
        while True:
            try:
                # 檢查任務是否仍然存在
                if task_id not in tasks:
                    yield f"data: {json.dumps({'error': '任務不存在或已過期'})}\n\n"
                    break
                
                # 檢查任務狀態
                task_status = tasks[task_id].get('status', 'unknown')
                if task_status == 'completed':
                    yield f"data: {json.dumps({'status': 'completed', 'message': '任務已完成'})}\n\n"
                    break
                elif task_status == 'error':
                    yield f"data: {json.dumps({'status': 'error', 'message': '任務執行失敗'})}\n\n"
                    break
                
                # 使用更短的超時時間，避免 Gunicorn 超時
                message_str = log_queue.get(timeout=10)  # 10秒超時
                if message_str == "<<TASK_DONE>>":
                    yield f"data: {json.dumps({'status': 'completed', 'message': '任務已完成'})}\n\n"
                    break
                
                # 確保傳送給前端的永遠是標準的 JSON 格式
                try:
                    # 嘗試解析，如果成功，表示它已經是 JSON 字串
                    json.loads(message_str)
                    yield f"data: {message_str}\n\n"
                except json.JSONDecodeError:
                    # 如果解析失敗，表示它是一個普通字串，我們將其包裝成 JSON
                    wrapped_message = json.dumps({"log": message_str})
                    yield f"data: {wrapped_message}\n\n"
                    
            except queue.Empty:
                current_time = time.time()
                # 每10秒發送一次心跳
                if current_time - last_heartbeat >= 10:
                    yield f"data: {json.dumps({'heartbeat': True, 'timestamp': current_time})}\n\n"
                    last_heartbeat = current_time
                    
                # 如果連接時間超過5分鐘，主動斷開
                if current_time - start_time > 3600:  # 60分鐘
                    yield f"data: {json.dumps({'error': '連接超時，請重新連接'})}\n\n"
                    break
    
    return Response(event_stream(), mimetype='text/event-stream')

@app.route('/api/results/<task_id>')
def get_results(task_id: str):
    """回傳指定任務的結果檔案列表"""
    results_dir = os.path.join(OUTPUT_FOLDER, task_id)
    if not os.path.isdir(results_dir):
        return jsonify({"error": "找不到結果目錄"}), 404

    files = []
    for filename in os.listdir(results_dir):
        files.append(filename)
    
    return jsonify(files)

@app.route('/api/single_result/<task_id>')
def get_single_result(task_id: str):
    """獲取單筆驗證的詳細結果"""
    try:
        # 檢查任務是否存在
        if task_id not in tasks:
            return jsonify({"error": "找不到指定的任務"}), 404
        
        # 直接從任務狀態中獲取結果
        if 'single_result' not in tasks[task_id]:
            app_logger.error(f"Task {task_id}: 任務狀態中沒有 single_result")
            return jsonify({"error": "任務尚未完成或結果不可用"}), 404
        
        single_result = tasks[task_id]['single_result']
        app_logger.info(f"Task {task_id}: 從任務狀態中獲取到結果: {single_result}")
        
        # 構建回傳結果
        result = {
            "task_id": task_id,
            "question": single_result.get("question", ""),
            "standard_answer": single_result.get("standard_answer", ""),
            "llm_response": single_result.get("llm_response", ""),
            "similarity_scores": single_result.get("similarity_scores", {
                "bert_score": 0,
                "cosine_similarity": 0
            })
        }
        
        app_logger.info(f"Task {task_id}: 回傳給前端的結果: {result}")
        return jsonify(result)
        
    except Exception as e:
        app_logger.error(f"獲取單筆結果時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": f"獲取結果時發生錯誤: {str(e)}"}), 500

@app.route('/api/download_single_result/<task_id>')
def download_single_result(task_id: str):
    """下載單筆驗證的詳細報告"""
    try:
        # 檢查任務是否存在
        if task_id not in tasks:
            return jsonify({"error": "找不到指定的任務"}), 404
        
        # 讀取 Excel 檔案
        # 先嘗試讀取原始檔案名稱
        excel_path = os.path.join(OUTPUT_FOLDER, task_id, f"{task_id}_single_verification.xlsx")
        if not os.path.exists(excel_path):
            # 如果找不到，嘗試讀取其他可能的檔案名稱
            task_dir = os.path.join(OUTPUT_FOLDER, task_id)
            if os.path.exists(task_dir):
                excel_files = [f for f in os.listdir(task_dir) if f.endswith('.xlsx')]
                if excel_files:
                    excel_path = os.path.join(task_dir, excel_files[0])
                else:
                    return jsonify({"error": "找不到結果檔案"}), 404
            else:
                return jsonify({"error": "找不到結果檔案"}), 404
        
        # 回傳 Excel 檔案供下載
        return send_from_directory(
            os.path.dirname(excel_path),
            os.path.basename(excel_path),
            as_attachment=True,
            download_name=f"single_verification_result_{task_id}.xlsx"
        )
        
    except Exception as e:
        app_logger.error(f"下載單筆結果時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": f"下載結果時發生錯誤: {str(e)}"}), 500

@app.route('/api/preview/<task_id>/<path:filename>')
def preview_file(task_id: str, filename: str):
    """根據檔案類型回傳預覽內容"""
    file_path = os.path.join(os.getcwd(), OUTPUT_FOLDER, task_id, filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "找不到檔案"}), 404

    try:
        file_ext = os.path.splitext(filename)[1].lower()
        
        if file_ext in ['.png', '.jpg', '.jpeg', '.gif']:
            # 對於圖片，直接回傳檔案的 URL 讓前端 <img> 標籤去載入
            # 注意：這裡回傳的是 HTML，而不是 JSON
            return f'<img src="/outputs/{task_id}/{filename}" style="max-width: 100%; height: auto;">'
        
        elif file_ext == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            # 將純文字包在 <pre> 標籤中以保留格式
            return f'<pre>{content}</pre>'

        elif file_ext in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path)
            # 將 DataFrame 轉換為 HTML 表格
            return df.to_html(classes='table table-striped table-hover', border=0, index=False)
        
        else:
            return '<p>不支援預覽此檔案類型。</p>'

    except Exception as e:
        app_logger.error(f"預覽檔案時發生錯誤: {e}", exc_info=True)
        return '<p style="color: red;">讀取檔案時發生錯誤。</p>', 500

@app.route('/outputs/<task_id>/<path:filename>')
def serve_output_file(task_id: str, filename: str):
    """安全地提供輸出目錄中的檔案"""
    directory = os.path.join(os.getcwd(), OUTPUT_FOLDER, task_id)
    return send_from_directory(directory, filename)

@app.route('/api/validate_connection', methods=['POST'])
def validate_connection():
    """API 端點，用於驗證與 AnythingLLM 的連線。"""
    data = request.json
    api_url = data.get('api_url')
    api_key = data.get('api_key')

    if not api_url or not api_key:
        return jsonify({"success": False, "message": "API URL 和金鑰為必填項"}), 400

    try:
        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }
        # 移除結尾的斜線以確保路徑正確
        validation_url = f"{api_url.rstrip('/')}/api/v1/auth"
        
        response = requests.get(validation_url, headers=headers, timeout=10)
        response.raise_for_status() # 如果狀態碼不是 2xx，則會引發 HTTPError
        
        return jsonify({"success": True, "message": "連線成功！"})

    except requests.exceptions.Timeout:
        return jsonify({"success": False, "message": "連線超時，請檢查 URL 和網路"}), 408
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            return jsonify({"success": False, "message": "認證失敗，API 金鑰無效"}), 401
        else:
            return jsonify({"success": False, "message": f"HTTP 錯誤: {e.response.status_code}"}), 500
    except requests.exceptions.RequestException as e:
        return jsonify({"success": False, "message": f"連線失敗，請檢查 API URL 是否正確"}), 500
    except Exception as e:
        app_logger.error(f"未知的驗證錯誤: {e}", exc_info=True)
        return jsonify({"success": False, "message": "發生未知的伺服器錯誤"}), 500


@app.route('/api/get_workspaces', methods=['POST'])
def get_workspaces():
    """API 端點，用於獲取 AnythingLLM 的工作區列表。"""
    data = request.json
    api_url = data.get('api_url')
    api_key = data.get('api_key')

    if not api_url or not api_key:
        return jsonify({"success": False, "message": "API URL 和金鑰為必填項"}), 400

    try:
        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }
        # 移除結尾的斜線以確保路徑正確
        workspaces_url = f"{api_url.rstrip('/')}/api/v1/workspaces"
        
        response = requests.get(workspaces_url, headers=headers, timeout=10)
        response.raise_for_status()
        
        workspaces_data = response.json()
        workspaces = workspaces_data.get('workspaces', [])
        
        # 格式化工作區列表，包含 id、name 和 slug
        formatted_workspaces = []
        for workspace in workspaces:
            formatted_workspaces.append({
                'id': workspace.get('id'),
                'name': workspace.get('name', '未命名工作區'),
                'slug': workspace.get('slug', '')
            })
        
        return jsonify({
            "success": True, 
            "workspaces": formatted_workspaces,
            "message": f"成功獲取 {len(formatted_workspaces)} 個工作區"
        })

    except requests.exceptions.Timeout:
        return jsonify({"success": False, "message": "連線超時，請檢查 URL 和網路"}), 408
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            return jsonify({"success": False, "message": "認證失敗，API 金鑰無效"}), 401
        elif e.response.status_code == 403:
            return jsonify({"success": False, "message": "權限不足，無法獲取工作區列表"}), 403
        else:
            return jsonify({"success": False, "message": f"HTTP 錯誤: {e.response.status_code}"}), 500
    except requests.exceptions.RequestException as e:
        return jsonify({"success": False, "message": f"連線失敗，請檢查 API URL 是否正確"}), 500
    except Exception as e:
        app_logger.error(f"獲取工作區列表時發生錯誤: {e}", exc_info=True)
        return jsonify({"success": False, "message": "發生未知的伺服器錯誤"}), 500


@app.route('/api/download_example')
def download_example():
    """API 端點，用於下載 Excel 範例檔案"""
    try:
        # 建立範例數據
        example_data = [
            ['什麼是 Python？', 'Python 是一種高級程式語言，以其簡潔的語法和強大的功能而聞名。它支援多種程式設計範式，包括程序式、物件導向和函數式程式設計。Python 被廣泛應用於網頁開發、數據分析、人工智慧、自動化腳本等領域。'],
            ['機器學習的主要類型有哪些？', '機器學習主要分為三種類型：1) 監督式學習：使用標記的訓練數據來學習輸入和輸出之間的映射關係，如分類和回歸問題。2) 非監督式學習：處理未標記的數據，發現數據中的隱藏模式，如聚類和降維。3) 強化學習：通過與環境互動來學習最佳策略，如遊戲AI和機器人控制。'],
            ['什麼是深度學習？', '深度學習是機器學習的一個子集，使用多層神經網路來學習數據的複雜模式。它能夠自動提取特徵，並在圖像識別、自然語言處理、語音識別等領域表現出色。深度學習的核心是多層感知機和各種神經網路架構，如卷積神經網路(CNN)、循環神經網路(RNN)和變壓器(Transformer)。'],
            ['什麼是 API？', 'API（應用程式介面）是一組定義了軟體組件如何相互通信的規則和協議。它允許不同的應用程式之間進行數據交換和功能調用。API 可以是 RESTful、GraphQL、SOAP 等不同類型，並提供標準化的方式來訪問服務和數據。'],
            ['什麼是資料庫？', '資料庫是一個有組織的數據集合，用於存儲、管理和檢索信息。常見的資料庫類型包括：1) 關聯式資料庫：如 MySQL、PostgreSQL，使用表格結構存儲數據。2) NoSQL 資料庫：如 MongoDB、Redis，支援更靈活的數據結構。3) 圖形資料庫：如 Neo4j，專門處理複雜的關係數據。'],
            ['什麼是容器化技術？', '容器化技術是一種虛擬化方法，將應用程式及其依賴項打包在一個標準化的單元中。Docker 是最流行的容器化平台，它提供了一致的運行環境，確保應用程式在不同環境中都能正常運行。容器化技術提高了部署效率、資源利用率和應用程式的可移植性。'],
            ['什麼是微服務架構？', '微服務架構是一種將應用程式分解為小型、獨立服務的設計模式。每個服務負責特定的業務功能，可以獨立開發、部署和擴展。微服務架構的優點包括：更好的可維護性、技術多樣性、獨立部署和故障隔離。但同時也帶來了分散式系統的複雜性挑戰。']
        ]

        # 建立 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "問答集範例"

        # 設定標題樣式
        center_alignment = Alignment(horizontal="center", vertical="center")

        # 寫入數據並設定樣式
        for row_idx, row_data in enumerate(example_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # 自動調整欄寬
                if col_idx == 1:  # 問題欄
                    ws.column_dimensions[cell.column_letter].width = 25
                elif col_idx == 2:  # 標準答案欄
                    ws.column_dimensions[cell.column_letter].width = 60
                else:  # 其他欄位
                    ws.column_dimensions[cell.column_letter].width = 15

        # 設定行高
        for row in range(1, len(example_data) + 1):
            if row == 1:
                ws.row_dimensions[row].height = 30
            else:
                ws.row_dimensions[row].height = 80

        # 將 Excel 檔案保存到記憶體
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # 回傳檔案
        return Response(
            excel_file.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=qa_example.xlsx'}
        )

    except Exception as e:
        app_logger.error(f"生成範例檔案時發生錯誤: {e}", exc_info=True)
        return jsonify({"error": "生成範例檔案時發生錯誤"}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5001) 